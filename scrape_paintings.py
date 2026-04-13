"""
Terraria Paintings Scraper
Scrapes https://terraria.wiki.gg/wiki/Paintings

Outputs:
  painting_images/  - PNG files from the wiki
  paintings.xlsx    - Excel workbook with T/F tracker + fraction counter
"""

import os
import re
import sys
import time
from io import BytesIO
from urllib.parse import urlparse, unquote

import requests
from bs4 import BeautifulSoup
from PIL import Image as PILImage
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_URL         = "https://terraria.wiki.gg"
WIKI_URL         = "https://terraria.wiki.gg/wiki/Paintings"
SCRIPT_DIR       = os.path.dirname(os.path.abspath(__file__))
IMAGES_DIR       = os.path.join(SCRIPT_DIR, "painting_images")
OUTPUT_XL        = os.path.join(SCRIPT_DIR, "paintings.xlsx")
HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; TerrariaWikiScraper/1.0)"}

# ── Excel layout constants ────────────────────────────────────────────────────
THUMB_W, THUMB_H = 48, 48
ROW_HEIGHT_PT    = 42
HEADER_ROW       = 2      # row with column labels
DATA_START_ROW   = 3      # first painting row

# Columns: (header text, column width in chars, horizontal alignment)
COLUMNS = [
    ("Obtained? (T/F)",     14,  "center"),  # A  type T = obtained, F or blank = not
    ("Painting",            9,   "center"),  # B  painting thumbnail
    ("Name",               28,   "left"),    # C
    ("Size",                9,   "center"),  # D
    ("Source / How to Get", 22,  "left"),    # E
    ("Description/Details", 50,  "left"),    # F
    ("Buy Price",          12,   "center"),  # G
    ("Sell Price",         12,   "center"),  # H
    ("Tooltip / Artist",   24,   "left"),    # I
    ("Placed Preview",      9,   "center"),  # J
]
LAST_COL = get_column_letter(len(COLUMNS))   # "J"

# Styles
HDR_FILL      = PatternFill("solid", fgColor="1F4E79")
HDR_FONT      = Font(bold=True, color="FFFFFF", size=11)
TITLE_FILL    = PatternFill("solid", fgColor="102A50")
TITLE_FONT    = Font(bold=True, color="FFFFFF", size=13)
ALT_FILL      = PatternFill("solid", fgColor="D6E4F0")
OBTAINED_FILL = PatternFill("solid", fgColor="B7E1CD")   # light green
BORDER_SIDE   = Side(style="thin", color="AAAAAA")
THIN_BORDER   = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                        top=BORDER_SIDE, bottom=BORDER_SIDE)
WRAP_ALIGN    = Alignment(wrap_text=True, vertical="center")
CENTER_ALIGN  = Alignment(horizontal="center", vertical="center")


# ─────────────────────────────────────────────────────────────────────────────
#  Fetching & image helpers
# ─────────────────────────────────────────────────────────────────────────────

def fetch_page(url: str) -> BeautifulSoup:
    print(f"Fetching {url} ...")
    resp = requests.get(url, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    return BeautifulSoup(resp.content, "html.parser", from_encoding="utf-8")


def full_image_url(src: str) -> str | None:
    if not src or src.startswith("data:"):
        return None
    if src.startswith("//"):
        return "https:" + src
    if src.startswith("/"):
        return BASE_URL + src
    return src


def clean_image_url(url: str) -> str:
    p = urlparse(url)
    return p._replace(query="").geturl()


def download_image(url: str, dest_dir: str, delay: float = 0.12) -> str | None:
    os.makedirs(dest_dir, exist_ok=True)
    filename = unquote(os.path.basename(urlparse(url).path))
    if not filename:
        return None
    filepath = os.path.join(dest_dir, filename)
    if os.path.exists(filepath):
        return filepath
    try:
        resp = requests.get(url, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        with open(filepath, "wb") as f:
            f.write(resp.content)
        time.sleep(delay)
        return filepath
    except Exception as exc:
        print(f"    Warning: could not download {url}: {exc}")
        return None


def make_thumbnail(image_path: str) -> BytesIO | None:
    try:
        with PILImage.open(image_path) as img:
            img = img.convert("RGBA")
            img.thumbnail((THUMB_W, THUMB_H), PILImage.LANCZOS)
            buf = BytesIO()
            img.save(buf, format="PNG")
            buf.seek(0)
            return buf
    except Exception as exc:
        print(f"    Warning: thumbnail failed for {image_path}: {exc}")
        return None


# ─────────────────────────────────────────────────────────────────────────────
#  Wiki parsing helpers
# ─────────────────────────────────────────────────────────────────────────────

def get_section_heading(table) -> str:
    for sib in table.find_all_previous(["h2", "h3"]):
        headline = sib.find(class_="mw-headline")
        text = (headline or sib).get_text(separator=" ", strip=True)
        if text:
            return text
    table_id = table.get("id", "")
    return (table_id.replace("-sell-table", "")
                    .replace("-table", "")
                    .replace("-", " ").title()) if table_id else "Unknown"


def map_columns(header_row) -> dict[str, int]:
    cols: dict[str, int] = {"painting": 0, "placed": 2}
    for i, cell in enumerate(header_row.find_all(["th", "td"])):
        text = cell.get_text(separator=" ", strip=True).lower()
        if i == 0:
            pass
        elif "name" in text:
            cols["name"] = i
        elif "placed" in text:
            cols["placed"] = i
        elif "size" in text or ("w" in text and "h" in text):
            cols["size"] = i
        elif "buy" in text:
            cols["buy"] = i
        elif "sell" in text:
            cols["sell"] = i
        elif "tooltip" in text:
            cols["tooltip"] = i
        elif any(k in text for k in ("desc", "obtain", "note", "source", "location")):
            cols["description"] = i
    return cols


def cell_text(cell) -> str:
    return re.sub(r"\s+", " ", cell.get_text(" ", strip=True)).strip()


def clean_name(raw: str) -> str:
    name = re.sub(r"\(Desktop[^)]*\)", "", raw)
    name = re.sub(r"Internal\s+Item\s+ID\s*:?\s*\d+", "", name, flags=re.IGNORECASE)
    return name.strip()


def get_cell(cells, cols, key) -> str:
    idx = cols.get(key)
    return cell_text(cells[idx]) if idx is not None and idx < len(cells) else ""


def get_img_src(cells, cols, key) -> str | None:
    idx = cols.get(key)
    if idx is None or idx >= len(cells):
        return None
    img = cells[idx].find("img")
    if not img:
        return None
    src = img.get("src", "")
    url = full_image_url(src)
    return clean_image_url(url) if url else None


def parse_row(cells, section: str, cols: dict) -> dict | None:
    if len(cells) < 4:
        return None

    name = ""
    name_idx = cols.get("name")
    if name_idx is not None and name_idx < len(cells):
        for span in cells[name_idx].find_all("span", title=True):
            t = span["title"].strip()
            if t and "versions" not in t.lower() and "item" not in t.lower():
                name = t
                break
        if not name:
            name = clean_name(cell_text(cells[name_idx]))
    if not name:
        paint_idx = cols.get("painting", 0)
        if paint_idx < len(cells):
            img = cells[paint_idx].find("img")
            if img:
                name = clean_name(img.get("alt", "").replace("_", " "))
    if not name:
        return None

    raw_size = get_cell(cells, cols, "size")
    size = re.sub(r"[\u00a0\s]+", " ", raw_size).strip()

    return {
        "name":               name,
        "size":               size,
        "how_to_obtain":      get_cell(cells, cols, "description"),
        "source_section":     section,
        "tooltip":            get_cell(cells, cols, "tooltip"),
        "buy_price":          get_cell(cells, cols, "buy"),
        "sell_price":         get_cell(cells, cols, "sell"),
        "painting_image_url": get_img_src(cells, cols, "painting"),
        "placed_image_url":   get_img_src(cells, cols, "placed"),
        "local_painting":     None,
        "local_placed":       None,
    }


# ─────────────────────────────────────────────────────────────────────────────
#  Scraping
# ─────────────────────────────────────────────────────────────────────────────

def scrape() -> list[dict]:
    soup = fetch_page(WIKI_URL)
    tables = soup.find_all("table", class_="Paintings-table")
    print(f"Found {len(tables)} painting tables\n")

    paintings: list[dict] = []
    seen: set[str] = set()

    for table in tables:
        section = get_section_heading(table)
        all_rows = table.find_all("tr")
        if not all_rows:
            continue
        cols = map_columns(all_rows[0])

        for row in all_rows[1:]:
            cells = row.find_all(["td", "th"])
            entry = parse_row(cells, section, cols)
            if entry is None or entry["name"] in seen:
                continue
            seen.add(entry["name"])

            if entry["painting_image_url"]:
                entry["local_painting"] = download_image(entry["painting_image_url"], IMAGES_DIR)
            if entry["placed_image_url"]:
                entry["local_placed"] = download_image(entry["placed_image_url"], IMAGES_DIR)

            img_file = (os.path.basename(entry["local_painting"])
                        if entry["local_painting"] else "(no image)")
            print(f"  [{section}] {entry['name']} | {entry['size']} | {img_file}")
            paintings.append(entry)

    return paintings


# ─────────────────────────────────────────────────────────────────────────────
#  Excel native checkbox injection (Excel 365 / Office 2024+)
# ─────────────────────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────────────────────
#  Excel export
# ─────────────────────────────────────────────────────────────────────────────

def write_excel(paintings: list[dict], path: str):
    print(f"\nBuilding Excel workbook ...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Paintings"

    n = len(paintings)
    last_data_row = DATA_START_ROW + n - 1
    data_range = f"A{DATA_START_ROW}:{LAST_COL}{last_data_row}"

    # ── Row 1: title + obtained counter ───────────────────────────────────────
    ws.merge_cells(f"A1:{LAST_COL}1")
    title_cell = ws["A1"]
    # Formula: live fraction that updates as user checks boxes
    title_cell.value = (
        '="Terraria Paintings Tracker  |  Obtained: "'
        f'&COUNTIF(A{DATA_START_ROW}:A{last_data_row},"T")'
        f'&" / "&COUNTA(C{DATA_START_ROW}:C{last_data_row})&" paintings"'
    )
    title_cell.font      = TITLE_FONT
    title_cell.fill      = TITLE_FILL
    title_cell.alignment = CENTER_ALIGN
    title_cell.border    = THIN_BORDER
    ws.row_dimensions[1].height = 26

    # ── Row 2: column headers ─────────────────────────────────────────────────
    for col_idx, (header, width, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=header)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.border    = THIN_BORDER
        cell.alignment = CENTER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[HEADER_ROW].height = 20

    # ── Conditional formatting: green row when A = "T" ───────────────────────
    obtained_rule = FormulaRule(
        formula=[f'$A{DATA_START_ROW}="T"'],
        fill=OBTAINED_FILL,
    )
    ws.conditional_formatting.add(data_range, obtained_rule)

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_offset, p in enumerate(paintings):
        row_idx  = DATA_START_ROW + row_offset
        is_alt   = (row_offset % 2 == 1)
        base_fill = ALT_FILL if is_alt else PatternFill()

        values = [
            "",                     # A  type T when obtained
            "",                     # B  painting image (inserted separately)
            p["name"],              # C
            p["size"],              # D
            p["source_section"],    # E
            p["how_to_obtain"],     # F
            p["buy_price"],         # G
            p["sell_price"],        # H
            p["tooltip"],           # I
            "",                     # J  placed image (inserted separately)
        ]

        for col_idx, val in enumerate(values, start=1):
            cell            = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border     = THIN_BORDER
            cell.fill       = base_fill
            _, _, align     = COLUMNS[col_idx - 1]
            cell.alignment  = Alignment(
                horizontal=align,
                vertical="center",
                wrap_text=(align == "left"),
            )

        ws.row_dimensions[row_idx].height = ROW_HEIGHT_PT

        # Embed painting thumbnail in col B (2)
        if p["local_painting"]:
            buf = make_thumbnail(p["local_painting"])
            if buf:
                xl_img         = XLImage(buf)
                xl_img.width   = THUMB_W
                xl_img.height  = THUMB_H
                ws.add_image(xl_img, f"B{row_idx}")

        # Embed placed thumbnail in col J (10)
        if p["local_placed"]:
            buf = make_thumbnail(p["local_placed"])
            if buf:
                xl_img         = XLImage(buf)
                xl_img.width   = THUMB_W
                xl_img.height  = THUMB_H
                ws.add_image(xl_img, f"J{row_idx}")

        if row_idx % 25 == 0:
            print(f"  ... wrote row {row_offset + 1} / {n}")

    # ── Freeze rows 1 & 2 ─────────────────────────────────────────────────────
    ws.freeze_panes = f"A{DATA_START_ROW}"

    # ── Auto-filter on header row ─────────────────────────────────────────────
    ws.auto_filter.ref = f"A{HEADER_ROW}:{LAST_COL}{last_data_row}"

    try:
        wb.save(path)
        print(f"Saved: {path}")
    except PermissionError:
        print(f"ERROR: Could not save {path} — is it open in Excel? Close it and try again.")
        raise



# ─────────────────────────────────────────────────────────────────────────────
#  Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    paintings = scrape()
    print(f"\nScraped {len(paintings)} unique paintings.")

    write_excel(paintings, OUTPUT_XL)

    print("\nDone!")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nCancelled.")
        sys.exit(0)
    except Exception as exc:
        print(f"\nERROR: {exc}")
        import traceback
        traceback.print_exc()
        input("\nPress Enter to close...")
        sys.exit(1)
